######################-----------------USER's MANUAL-------------------------########################### 
#CAUTION: 1.to run this program, intialize an empty xlsx file in your target path,
#         name it "Finished.xlsx"
#         Then run this code, the browser will automatically pop up
#         After you login, jump to the supply page
#         Then move back to program and press enter to fatch the suborder id
#
#         2. If you are running this program first time on your computer, please manually query by order id once.
#           Then tap F12 and in the check the most recent request in the network interface, and find the token in the header in this request.
#           Then copy its token into the request module which is labled "REQUEST" in this program
#
#############################################################################################################
##############################INTRODUCTION##################################################################
#This program automatically fatches the sub-order ID for multi-item orders
#Since the douyindianpu lacks the function of pairing sub-order ID and Main order ID,
#It requires user to pair the information by querying the information of order with manual operations
#These program simplifes this process by automatically sending query request and fatch data of the order
#The progrom is based on the REST API of the web. 
#The process of this program is as follow:
#read the Prime ID of the whole order -> send out query request to obtain the information of the sub orders
# -> obtain the JSON data package -> obtain the item name and the sub-order ID
#############################################################################################################
#############################END OF PROGRAM INTRODUCTION#####################################################
#############################################################################################################

#####BASIC IDEA:
'''
Implementation:
The information of the order IDs need querying is stored in a XLXS (excel) file
The program read the data into a list and automatically modify the query request by changing the orderID string in the request
The the requested package is obtained by the program.
The data package is a JSON file. The program use the .json method to analyze the data package
The basic data structure of this package is a combinition of arrays and dictionaries
The information we want(sub-order ID, item name)is stored in following path:
Item Name:
FatchedPackage['data'][0]['children']['Sub-Order Index']['product']['name']
Sub-Order ID:
FatchedPackage['data'][0]['children']['Sub-Order Index']['order']['order_id']

The number of sub-orders can be obtained by len() method on the FatchedPackage['data'][0]['children']
The Sub-Order Index is an integer starting from 0

After analyzing the obtained data, the information of sub-ID and item name will be stored into a xlsx file
The format of the output xlsx file is:

MAIN ID|Sub-ID|Item Name
'''
##########################################END OF MANUAL################################################################
##########################################PROGRAM STARTS FROM HERE#####################################################
import xlrd
import openpyxl
import seleniumrequests
import time
import random
import tkinter
from tkinter import filedialog
from seleniumrequests import Chrome

#ReadMainID function reads all the main IDs into a list
def ReadMainID(path):
    Main_IDs = []

    #This function use the xlrd library to read xlsx
    workbook_read = xlrd.open_workbook(path)
    worksheet = workbook_read.sheet_by_index(0)
    nRows = worksheet.nrows
    #Use loop to store the information into memory
    i = 0
    while i <  nRows:
        Main_IDs.append(worksheet.cell(i,0).value) 
        i += 1
    print(Main_IDs)
    return Main_IDs

#RequestID completes the query order ID
def RequestID(mainids, work_sheetw, driver):
    row_index = 1
    index_id = 0

    #This loop makes up the main process of sending querying request
    while index_id < len(mainids):
        #For one main order ID, send out the request once
        #response is the fatched data package
        #Use selenium-request to send out request
        #The order_id section is variable, by changing it we query information of different order_id

        baset = str(1594003540704)
        ######### REQUEST ##########
        response = webdriver.request('GET', 'https://fxg.jinritemai.com/order/torder/searchlist?order_id=' + mainids[index_id] + '&order_status=&final_status=0&after_sale_status=&c_type=&pay_type=&product_name=&logistics_id=&logistics_code=&post_receiver=&post_tel=&start_time=&end_time=&start_receipt_time=&end_receipt_time=&timeout=0&badge=0&urge_tag=0&page=0&pageSize=20&total=1&b_type=&order_supply_type=&order_type_sub=-1&order_type=-1&is_ad=-1&c_biz=&sub_shop_ids=&order=create_time&is_desc=desc&__t='+baset+'&__token=a96b2f53dea5f99b18546cee73ef918c')
        ItemAmount = len(response.json()['data'][0]['children']) #Number of suborders

        #Check whether the order has sub_orders
        if ItemAmount == 1:
            index_sub +=1 #move to next sub-order
            row_index +=1 #target next row in the xlsx
            continue
            #If this is a single order than skip this loop
        else:
            index_sub = 0
            #This loop write the infofmation of the suborders
            while index_sub <  ItemAmount:
                #Targeting the cells in xlsx
                Main_ID_Cell = work_sheetw.cell(row = row_index, column = 1)
                Sub_ID_Cell = work_sheetw.cell(row = row_index, column = 2)
                ItemName_Cell = work_sheetw.cell(row = row_index, column = 3)

                #Write data into the cells
                Main_ID_Cell.value = mainids[index_id]
                Sub_ID_Cell.value = response.json()['data'][0]['children'][index_sub]['order']['order_id']
                ItemName_Cell.value = response.json()['data'][0]['children'][index_sub]['product']['name']

                #Print data out in termianl to check
                # numberofrow--MainID--SubID--ItemName
                # INFORMATION IN TERMINAL MAY BE INCOMPLETE
                print(str(row_index) + ' ' + mainids[index_id] + '  ' +
                response.json()['data'][0]['children'][index_sub]['order']['order_id']+'  '+
                response.json()['data'][0]['children'][index_sub]['product']['name'])
                ###END OF THE MAIN BODY OF THE LOOP
                time.sleep(1)
                #Every time the program record a line in the xlsx file, pause for 1 second
                #thus for each query, there is a variable delay for next query
                #Use this method to hide from the detection of robot

                index_sub +=1 #move to next sub-order
                row_index +=1 #target next row in the xlsx
            
            if index_id%25 == 0:
                time.sleep(5)
                baset = str(1594003540704 + index_id)
########
#The request_time section in the query request is also variable, however its function is unknown
#Guessing it's use for the detection of robot
#Therefore, change its value for every several request
###################
        #time.sleep(1)
        index_id += 1



if __name__ == "__main__":
    #Enter the path of the file storing the information of the main order id
    path = filedialog.askdirectory() + '\\'
    #Filename_read = str(input('ENTER THE FILENAME: '))
    filepath_full = filedialog.askopenfilename()

    listIDs = ReadMainID(filepath_full)
    #Open the file for temp storing
    workbook_write = openpyxl.load_workbook(path + 'Finished.xlsx')
    worksheet_write = workbook_write.worksheets[0]
    #Set browser as Chrome
    webdriver = Chrome()
    #Open the login website
    webdriver.get("https://open.snssdk.com/oauth/authorize/?client_key=ttae0f96cae89a91&state=toutiao&response_type=code&scope=mobile%2Cuser_info&redirect_uri=https%3A%2F%2Ffxg.jinritemai.com%2Findex.html%23%2Fffa%2Flogin")
    
    n = input("Press Enter")#Use an empty input to pause the program and wait for the finsh of manual log in
    start = time.time()
    RequestID(listIDs, worksheet_write, webdriver)
    end = time.time()
    workbook_write.save(path+'Temp.xlsx')
    print('SAVE DONE')
    print(end- start)
